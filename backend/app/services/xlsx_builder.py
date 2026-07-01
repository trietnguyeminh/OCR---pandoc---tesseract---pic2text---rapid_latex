"""PDF -> Excel (faithful to the Markdown the pipeline produced).

Markdown tables become real Excel columns. CONSECUTIVE tables that share the
same header row are merged into ONE sheet (so a roster split across many pages
becomes a single continuous table). Rows are aligned to the header so cells
never shift. Any remaining free text goes to a "Nội dung" sheet.
"""
from __future__ import annotations

import logging
import re

logger = logging.getLogger("formudoc.xlsx")

_PAGEBREAK = re.compile(r"```\{=openxml\}.*?```", re.S)
_RULE = re.compile(r"^\|?[\s:|_-]{3,}\|?$")


def _clean_line(line: str) -> str:
    line = re.sub(r"^\s{0,3}#{1,6}\s*", "", line)
    line = re.sub(r"</?span[^>]*>", "", line)
    return line.strip()


def _is_table_line(line: str) -> bool:
    s = line.strip()
    return s.startswith("|") and s.count("|") >= 2


def _split_row(line: str) -> list:
    s = line.strip()
    if s.startswith("|"):
        s = s[1:]
    if s.endswith("|"):
        s = s[:-1]
    return [_clean_line(c) for c in s.split("|")]


def _is_sep_cells(cells: list) -> bool:
    return bool(cells) and all(re.fullmatch(r":?-{2,}:?", c or "") for c in cells)


def _hdr_key(header: list) -> tuple:
    """Normalised header signature so continuation tables with the same columns
    (possibly minor spacing/case differences) are recognised as the same table."""
    return tuple(re.sub(r"[\s]+", "", (c or "").lower()) for c in header)


def _align(row: list, ncols: int) -> list:
    if len(row) > ncols:                     # extra cells -> merge into the last
        return row[:ncols - 1] + [" ".join(x for x in row[ncols - 1:] if x)]
    if len(row) < ncols:                     # short row -> pad
        return row + [""] * (ncols - len(row))
    return list(row)


def build_xlsx_from_markdown(md: str, out_path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    hdr_fill = PatternFill("solid", fgColor="6D5EFC")

    def style_header(ws):
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = hdr_fill
            c.alignment = Alignment(vertical="center", wrap_text=True)

    tables = []          # each = list of rows (row = list of cells), row0 = header
    text_rows = []

    for pno, part in enumerate(_PAGEBREAK.split(md or ""), start=1):
        lines = part.split("\n")
        i = 0
        while i < len(lines):
            if _is_table_line(lines[i]):
                block = []
                while i < len(lines) and _is_table_line(lines[i]):
                    cells = _split_row(lines[i])
                    if not _is_sep_cells(cells):
                        block.append(cells)
                    i += 1
                if len(block) >= 2:
                    tables.append(block)
                elif block:
                    text_rows.append((pno, " | ".join(block[0])))
            else:
                t = _clean_line(lines[i])
                if t and t != "---" and not t.startswith("```") and not _RULE.match(t):
                    text_rows.append((pno, t))
                i += 1

    # --- merge CONSECUTIVE tables that share the same header ------------- #
    groups = []          # each = {"header": [...], "rows": [...]}
    for block in tables:
        header, data = block[0], block[1:]
        key = _hdr_key(header)
        if groups and groups[-1]["key"] == key:
            groups[-1]["rows"].extend(data)             # continuation page
        else:
            groups.append({"key": key, "header": header, "rows": list(data)})

    # biggest table first so the main roster is the sheet Excel opens on
    groups.sort(key=lambda g: len(g["rows"]), reverse=True)

    # --- write each merged table to its own sheet ----------------------- #
    for idx, g in enumerate(groups, start=1):
        ws = wb.create_sheet(title=("Bảng %d" % idx)[:31])
        ncols = max(1, len(g["header"]))
        ws.append(_align(g["header"], ncols))
        style_header(ws)
        for r in g["rows"]:
            ws.append(_align(r, ncols))
        for c in range(1, ncols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 20
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"

    # --- free text -> "Nội dung" sheet ---------------------------------- #
    if text_rows:
        ws = wb.create_sheet(title="Nội dung")
        ws.append(["Trang", "Nội dung"])
        style_header(ws)
        for pno, t in text_rows:
            ws.append([pno, t])
        ws.column_dimensions["A"].width = 7
        ws.column_dimensions["B"].width = 110
        for row in ws.iter_rows(min_row=2):
            row[0].alignment = Alignment(horizontal="center", vertical="top")
            row[1].alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"

    if not wb.sheetnames:
        ws = wb.create_sheet(title="Nội dung")
        ws.append(["Trang", "Nội dung"])
        style_header(ws)

    if wb.sheetnames:
        wb.active = 0
    wb.save(out_path)
    logger.info("xlsx built: %s (%d table sheet(s) from %d table block(s), %d text row(s))",
                out_path, len(groups), len(tables), len(text_rows))
    return out_path
